#!/usr/bin/env python3
"""
Build kanji_data.json for the kanji study PWA.

Pipeline:
  1. Extract ~2136 jouyou kanji from 常用漢字.xlsx (onyomi-grouped)
  2. Enrich with KANJIDIC2 (meanings, readings)
  3. Find common compound words from JMdict
  4. Output kanji_data.json
"""

import json
import gzip
import os
import re
import sys
import urllib.request
import xml.etree.ElementTree as ET
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
CACHE_DIR = BASE_DIR / "cache"
XLSX_PATH = Path.home() / "Downloads" / "常用漢字.xlsx"
OUTPUT_PATH = BASE_DIR / "kanji_data.json"

KANJIDIC2_URL = "http://www.edrdg.org/kanjidic/kanjidic2.xml.gz"
KANJIDIC2_GZ = CACHE_DIR / "kanjidic2.xml.gz"
KANJIDIC2_XML = CACHE_DIR / "kanjidic2.xml"

JMDICT_URL = "http://ftp.edrdg.org/pub/Nihongo/JMdict_e.gz"
JMDICT_GZ = CACHE_DIR / "JMdict_e.gz"
JMDICT_XML = CACHE_DIR / "JMdict_e.xml"

MAX_DISPLAY_WORDS = 3
MAX_CANDIDATES = 30


# ---------------------------------------------------------------------------
# Step 1: Extract kanji from xlsx
# ---------------------------------------------------------------------------
def extract_kanji_from_xlsx():
    """Extract kanji grouped by onyomi from the spreadsheet."""
    import openpyxl

    print("Step 1: Extracting kanji from xlsx...")
    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb["漢字"]

    # Find all onyomi header columns (row 2, non-None, col >= 2)
    onyomi_cols = []
    for col in range(2, ws.max_column + 1):
        v = ws.cell(row=2, column=col).value
        if v is not None:
            onyomi_cols.append((col, str(v).strip()))

    kanji_list = []
    seen = set()

    for col, onyomi_label in onyomi_cols:
        for row in range(3, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            raw = str(v).strip()
            if not raw:
                continue
            # Strip parenthetical/bracket content: 亜（亞） → 亜, 貝  (かい) → 貝, 餌 ［餌］ → 餌
            kanji_char = re.sub(r"\s*[（(\[［].*?[）)\]］]", "", raw).strip()
            if len(kanji_char) != 1:
                print(f"  Warning: unexpected value after cleanup: {repr(raw)} -> {repr(kanji_char)}")
                continue
            if kanji_char in seen:
                print(f"  Warning: duplicate kanji {kanji_char} in col {col} ({onyomi_label})")
                continue
            seen.add(kanji_char)
            kanji_list.append({
                "kanji": kanji_char,
                "onyomiGroup": onyomi_label,
            })

    wb.close()
    print(f"  Extracted {len(kanji_list)} kanji from {len(onyomi_cols)} onyomi groups.")
    return kanji_list


# ---------------------------------------------------------------------------
# Step 2: Download and parse KANJIDIC2
# ---------------------------------------------------------------------------
def download_file(url, dest_gz, dest_xml):
    """Download a gzipped file and decompress it, with caching."""
    if dest_xml.exists():
        print(f"  Cached: {dest_xml.name}")
        return

    if not dest_gz.exists():
        print(f"  Downloading {url} ...")
        try:
            urllib.request.urlretrieve(url, str(dest_gz))
        except Exception as e:
            print(f"  ERROR downloading {url}: {e}", file=sys.stderr)
            sys.exit(1)
        print(f"  Downloaded: {dest_gz.name} ({dest_gz.stat().st_size / 1024 / 1024:.1f} MB)")
    else:
        print(f"  Cached: {dest_gz.name}")

    print(f"  Decompressing {dest_gz.name} ...")
    with gzip.open(str(dest_gz), "rb") as f_in:
        with open(str(dest_xml), "wb") as f_out:
            while True:
                chunk = f_in.read(1024 * 1024)
                if not chunk:
                    break
                f_out.write(chunk)
    print(f"  Decompressed: {dest_xml.name} ({dest_xml.stat().st_size / 1024 / 1024:.1f} MB)")


def parse_kanjidic2(kanji_set):
    """Parse KANJIDIC2 XML and return a dict mapping kanji -> info."""
    print("Step 2: Parsing KANJIDIC2...")
    download_file(KANJIDIC2_URL, KANJIDIC2_GZ, KANJIDIC2_XML)

    kd = {}
    # Use iterparse to handle large XML efficiently
    context = ET.iterparse(str(KANJIDIC2_XML), events=("end",))
    for event, elem in context:
        if elem.tag != "character":
            continue

        literal = elem.findtext("literal")
        if literal not in kanji_set:
            elem.clear()
            continue

        meanings = []
        onyomi = []
        kunyomi = []

        # readings and meanings are in reading_meaning group
        rmgroup = elem.find(".//rmgroup")
        if rmgroup is not None:
            for reading in rmgroup.findall("reading"):
                r_type = reading.get("r_type")
                if r_type == "ja_on":
                    onyomi.append(reading.text)
                elif r_type == "ja_kun":
                    kunyomi.append(reading.text)

            for meaning in rmgroup.findall("meaning"):
                # Only English meanings (no m_lang attribute = English)
                if meaning.get("m_lang") is None:
                    meanings.append(meaning.text)

        kd[literal] = {
            "meanings": meanings,
            "onyomi": onyomi,
            "kunyomi": kunyomi,
        }
        elem.clear()

    print(f"  Parsed data for {len(kd)} kanji.")
    return kd


# ---------------------------------------------------------------------------
# Step 3: Parse JMdict for common words
# ---------------------------------------------------------------------------
def parse_jmdict(kanji_set):
    """Parse JMdict XML to find common compound words for each kanji.

    Returns:
      - words dict: kanji -> list of top display words
      - reading_evidence dict: kanji -> list of (word, reading) for reading validation
    """
    print("Step 3: Parsing JMdict for common words...")
    download_file(JMDICT_URL, JMDICT_GZ, JMDICT_XML)

    # Priority tags that indicate common/important words
    PRIORITY_TAGS = {"news1", "ichi1", "spec1", "ichi2", "news2", "spec2"}

    # Build per-kanji word lists
    kanji_words = {k: [] for k in kanji_set}
    # Track (word, reading) pairs per kanji for reading validation
    kanji_reading_evidence = {k: [] for k in kanji_set}

    entry_count = 0
    matched_count = 0

    context = ET.iterparse(str(JMDICT_XML), events=("end",))
    for event, elem in context:
        if elem.tag != "entry":
            continue
        entry_count += 1

        # Get kanji elements (k_ele)
        k_eles = elem.findall("k_ele")
        if not k_eles:
            elem.clear()
            continue

        # Get reading elements (r_ele)
        r_eles = elem.findall("r_ele")

        # Get sense/meaning elements
        senses = elem.findall("sense")

        for k_ele in k_eles:
            keb = k_ele.findtext("keb")  # kanji form of the word
            if not keb:
                continue

            # Get priority from kanji element (ke_pri)
            ke_pris = [p.text for p in k_ele.findall("ke_pri")]

            # Get matching reading and its priority (re_pri)
            reading = None
            re_pris = []
            for r_ele in r_eles:
                re_restr = [r.text for r in r_ele.findall("re_restr")]
                if re_restr and keb not in re_restr:
                    continue
                reading = r_ele.findtext("reb")
                re_pris = [p.text for p in r_ele.findall("re_pri")]
                break

            if not reading:
                continue

            # Score from ke_pri (kanji form is standard)
            # nf values = frequency rank (lower = more common), use directly
            # PRIORITY_TAGS without nf = common but unranked, default to 5
            ke_has = False
            ke_nf = None
            for pri in ke_pris:
                if pri and pri.startswith("nf"):
                    ke_has = True
                    try:
                        nf = int(pri[2:])
                        ke_nf = min(ke_nf, nf) if ke_nf is not None else nf
                    except ValueError:
                        pass
                elif pri in PRIORITY_TAGS:
                    ke_has = True
            ke_score = ke_nf if ke_nf is not None else (5 if ke_has else 100)

            # Score from re_pri (word common but kanji form less standard)
            re_has = False
            re_nf = None
            if not ke_has:
                for pri in re_pris:
                    if pri and pri.startswith("nf"):
                        re_has = True
                        try:
                            nf = int(pri[2:])
                            re_nf = min(re_nf, nf) if re_nf is not None else nf
                        except ValueError:
                            pass
                    elif pri in PRIORITY_TAGS:
                        re_has = True
            re_score = (re_nf + 3 if re_nf is not None else 8) if re_has else 100

            has_priority = ke_has or re_has
            if not has_priority:
                continue

            score = min(ke_score, re_score)

            # Reading evidence: only from ke_pri words (strict validation)
            if ke_has:
                for ch in keb:
                    if ch in kanji_reading_evidence:
                        kanji_reading_evidence[ch].append((keb, reading))

            # Get first English meaning
            meaning_parts = []
            if senses:
                first_sense = senses[0]
                for gloss in first_sense.findall("gloss"):
                    lang = gloss.get("{http://www.w3.org/XML/1998/namespace}lang", "eng")
                    if lang == "eng" and gloss.text:
                        meaning_parts.append(gloss.text)
            meaning = "; ".join(meaning_parts[:3]) if meaning_parts else None

            if not meaning:
                continue

            # Slight preference for 2-char compounds, minimal penalty otherwise
            if len(keb) == 1:
                length_penalty = 2
            else:
                length_penalty = max(0, len(keb) - 2)
            adjusted_score = score + length_penalty

            for ch in keb:
                if ch in kanji_words:
                    kanji_words[ch].append((adjusted_score, keb, reading, meaning))
                    matched_count += 1

        elem.clear()

    print(f"  Processed {entry_count} JMdict entries, {matched_count} word-kanji associations.")

    # Sort and keep top candidates per kanji (selection happens later)
    result = {}
    for k, words in kanji_words.items():
        if not words:
            result[k] = []
            continue
        words.sort(key=lambda x: (x[0], len(x[1])))
        seen = set()
        unique = []
        for score, word, reading, meaning in words:
            pair = (word, reading)
            if pair in seen:
                continue
            seen.add(pair)
            unique.append({"word": word, "reading": reading, "meaning": meaning})
            if len(unique) >= MAX_CANDIDATES:
                break
        result[k] = unique

    kanji_with_words = sum(1 for v in result.values() if v)
    print(f"  Found words for {kanji_with_words}/{len(kanji_set)} kanji.")
    return result, kanji_reading_evidence


# ---------------------------------------------------------------------------
# Step 3.5: Filter readings by actual usage in common words
# ---------------------------------------------------------------------------
def kata_to_hira(s):
    """Convert katakana string to hiragana."""
    return ''.join(
        chr(ord(c) - 0x60) if '\u30A1' <= c <= '\u30F6' else c
        for c in s
    )


def reading_appears_in(hira, word_readings):
    """Check if a hiragana reading appears in any common word's reading.

    Also handles gemination (e.g. がく matching がっこう via がっ).
    """
    for _, wr in word_readings:
        if hira in wr:
            return True
        # Gemination: last mora becomes っ (e.g. ガク→がっ in 学校/がっこう)
        if len(hira) >= 2 and hira[:-1] + 'っ' in wr:
            return True
    return False


def filter_readings(kanjidic, kanji_reading_evidence):
    """Remove readings from KANJIDIC2 that don't appear in any common word."""
    print("Step 3.5: Filtering readings by usage in common words...")

    stats = {'on_kept': 0, 'on_dropped': 0, 'kun_kept': 0, 'kun_dropped': 0}

    for kanji, info in kanjidic.items():
        evidence = kanji_reading_evidence.get(kanji, [])
        if not evidence:
            # No common words at all — keep readings as-is (better than nothing)
            stats['on_kept'] += len(info['onyomi'])
            stats['kun_kept'] += len(info['kunyomi'])
            continue

        # Filter onyomi
        filtered_on = []
        for on in info['onyomi']:
            hira = kata_to_hira(on)
            if reading_appears_in(hira, evidence):
                filtered_on.append(on)
                stats['on_kept'] += 1
            else:
                stats['on_dropped'] += 1

        # Filter kunyomi — use stem (before the . okurigana marker)
        filtered_kun = []
        for kun in info['kunyomi']:
            stem = kun.split('.')[0].lstrip('-') if '.' in kun else kun.lstrip('-')
            if stem and reading_appears_in(stem, evidence):
                filtered_kun.append(kun)
                stats['kun_kept'] += 1
            else:
                stats['kun_dropped'] += 1

        info['onyomi'] = filtered_on
        info['kunyomi'] = filtered_kun

    print(f"  Onyomi:  kept {stats['on_kept']}, dropped {stats['on_dropped']}")
    print(f"  Kunyomi: kept {stats['kun_kept']}, dropped {stats['kun_dropped']}")


# ---------------------------------------------------------------------------
# Step 3.75: Select display words that cover readings
# ---------------------------------------------------------------------------
def select_display_words(onyomi, kunyomi, candidates):
    """Pick up to MAX_DISPLAY_WORDS that best represent the kanji's readings.

    Phase 1: For each reading, pick the most common word that uses it.
    Phase 2: Fill remaining slots with the most common words overall.
    """
    if not candidates:
        return []

    selected = []
    used = set()

    # Build list of readings to cover (hiragana stems)
    readings_to_cover = []
    for on in onyomi:
        readings_to_cover.append(kata_to_hira(on))
    for kun in kunyomi:
        stem = kun.split('.')[0].lstrip('-') if '.' in kun else kun.lstrip('-')
        if stem:
            readings_to_cover.append(stem)

    # Phase 1: Cover each reading with the best available word
    for hira in readings_to_cover:
        if len(selected) >= MAX_DISPLAY_WORDS:
            break
        for c in candidates:
            if c['word'] in used:
                continue
            # Check if this word demonstrates this reading
            if reading_appears_in(hira, [(c['word'], c['reading'])]):
                selected.append(c)
                used.add(c['word'])
                break

    # Phase 2: Fill remaining slots with most common words
    for c in candidates:
        if len(selected) >= MAX_DISPLAY_WORDS:
            break
        if c['word'] not in used:
            selected.append(c)
            used.add(c['word'])

    return selected


# ---------------------------------------------------------------------------
# Step 4: Assemble and output JSON
# ---------------------------------------------------------------------------
def main():
    CACHE_DIR.mkdir(parents=True, exist_ok=True)

    # Step 1: Extract kanji
    kanji_list = extract_kanji_from_xlsx()
    kanji_set = {entry["kanji"] for entry in kanji_list}

    # Step 2: Enrich with KANJIDIC2
    kanjidic = parse_kanjidic2(kanji_set)

    # Step 3: Get common words from JMdict
    jmdict_words, kanji_reading_evidence = parse_jmdict(kanji_set)

    # Step 3.5: Filter readings to only those used in common words
    filter_readings(kanjidic, kanji_reading_evidence)

    # Step 4: Assemble final data
    print("Step 4: Assembling output...")
    output = []
    missing_kd = 0
    for entry in kanji_list:
        k = entry["kanji"]
        kd_info = kanjidic.get(k, {})
        if not kd_info:
            missing_kd += 1

        on = kd_info.get("onyomi", [])
        kun = kd_info.get("kunyomi", [])
        words = select_display_words(on, kun, jmdict_words.get(k, []))

        output.append({
            "kanji": k,
            "onyomiGroup": entry["onyomiGroup"],
            "meanings": kd_info.get("meanings", []),
            "onyomi": on,
            "kunyomi": kun,
            "words": words,
        })

    if missing_kd:
        print(f"  Note: {missing_kd} kanji not found in KANJIDIC2.")

    # Fallback: find words for kanji with none (no priority requirement)
    empty_kanji = {e["kanji"] for e in output if not e["words"]}
    if empty_kanji:
        print(f"  Finding fallback words for {len(empty_kanji)} kanji...")
        fallback = {k: [] for k in empty_kanji}
        context = ET.iterparse(str(JMDICT_XML), events=("end",))
        for _, elem in context:
            if elem.tag != "entry":
                continue
            k_eles = elem.findall("k_ele")
            r_eles = elem.findall("r_ele")
            senses = elem.findall("sense")
            for k_ele in k_eles:
                keb = k_ele.findtext("keb")
                if not keb:
                    continue
                targets = [ch for ch in keb if ch in empty_kanji and len(fallback[ch]) < 3]
                if not targets:
                    continue
                reading = None
                for r_ele in r_eles:
                    re_restr = [r.text for r in r_ele.findall("re_restr")]
                    if re_restr and keb not in re_restr:
                        continue
                    reading = r_ele.findtext("reb")
                    break
                if not reading:
                    continue
                meaning_parts = []
                if senses:
                    for gloss in senses[0].findall("gloss"):
                        lang = gloss.get("{http://www.w3.org/XML/1998/namespace}lang", "eng")
                        if lang == "eng" and gloss.text:
                            meaning_parts.append(gloss.text)
                meaning = "; ".join(meaning_parts[:3]) if meaning_parts else None
                if not meaning:
                    continue
                for ch in targets:
                    fallback[ch].append({"word": keb, "reading": reading, "meaning": meaning})
            elem.clear()
            if all(len(v) >= 3 for v in fallback.values()):
                break
        for entry in output:
            if not entry["words"] and fallback.get(entry["kanji"]):
                entry["words"] = fallback[entry["kanji"]]
        found = sum(1 for v in fallback.values() if v)
        print(f"  Found fallback words for {found}/{len(empty_kanji)} kanji.")

    # Write output
    with open(str(OUTPUT_PATH), "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    file_size = OUTPUT_PATH.stat().st_size / 1024
    print(f"\nDone! Wrote {len(output)} kanji entries to {OUTPUT_PATH}")
    print(f"  File size: {file_size:.1f} KB")

    # Quick stats
    with_meanings = sum(1 for e in output if e["meanings"])
    with_words = sum(1 for e in output if e["words"])
    total_words = sum(len(e["words"]) for e in output)
    print(f"  Kanji with meanings: {with_meanings}")
    print(f"  Kanji with words: {with_words}")
    print(f"  Total word entries: {total_words}")
    print(f"  First entry: {json.dumps(output[0], ensure_ascii=False)}")
    print(f"  Last entry: {json.dumps(output[-1], ensure_ascii=False)}")


if __name__ == "__main__":
    main()
